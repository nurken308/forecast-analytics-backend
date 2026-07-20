from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from db.database import get_session
from repositories.ezmes_repository import EzmesRepository
from service.daily_forecast_service import (
    build_daily_forecast,
    get_current_daily_forecast,
    get_daily_forecast_history,
    auto_update_daily_forecast,
    force_recalculate_daily_forecast,
)


router = APIRouter(
    prefix="/daily-forecast",
    tags=["Daily Forecast"],
)


DAILY_EXPORT_SEGMENTS = [
    ("retail", "Розница"),
    ("bk", "БК"),
    ("bk_soft", "БК_СОФТ"),
    ("vkl", "ВКЛ"),
    ("vkl_soft", "ВКЛ_СОФТ"),
    ("rb_z_ip", "РБ З ИП"),
    ("rb_z_ip_soft", "РБ З ИП(с)"),
    ("rb_z_too", "РБ З ТОО"),
    ("rb_bz_ip", "РБ БЗ ИП"),
    ("rb_bz_too", "РБ БЗ ТОО"),
]


def serialize_value(value):
    return {
        "id": value.id,
        "period_date": value.period_month,
        "forecast_value": (
            float(value.forecast_value)
            if value.forecast_value is not None
            else None
        ),
        "fact_value": (
            float(value.fact_value)
            if value.fact_value is not None
            else None
        ),
        "abs_error": (
            float(value.abs_error)
            if value.abs_error is not None
            else None
        ),
        "pct_error": (
            float(value.pct_error)
            if value.pct_error is not None
            else None
        ),
        "status": value.status,
    }


def serialize_run(run):
    return {
        "id": run.id,
        "forecast_type": run.forecast_type,
        "segment_id": run.segment_id,
        "model_name": run.model_name,
        "status": run.status,
        "parent_run_id": run.parent_run_id,
        "created_at": run.created_at,
        "updated_at": run.updated_at,
        "values": [
            serialize_value(value)
            for value in sorted(
                run.values,
                key=lambda item: item.period_month,
            )
        ],
    }


def format_excel_sheet(
    worksheet,
    sheet_title: str,
    values,
):
    """
    Заполняет один лист:
    - таблица;
    - форматирование;
    - встроенный линейный график Факт / Прогноз.
    """

    headers = [
        "Период",
        "Факт",
        "Прогноз",
        "Отклонения",
    ]

    worksheet.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = thin_border

    sorted_values = sorted(
        values,
        key=lambda item: item.period_month,
    )

    for value in sorted_values:
        period_date = datetime.strptime(
            value.period_month,
            "%Y-%m-%d",
        )

        fact_value = (
            float(value.fact_value)
            if value.fact_value is not None
            else None
        )

        forecast_value = (
            float(value.forecast_value)
            if value.forecast_value is not None
            else None
        )

        deviation = (
            fact_value - forecast_value
            if (
                fact_value is not None
                and forecast_value is not None
            )
            else None
        )

        worksheet.append(
            [
                period_date,
                fact_value,
                forecast_value,
                deviation,
            ]
        )

    last_row = worksheet.max_row

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:D{last_row}"

    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 18

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=last_row,
        min_col=1,
        max_col=4,
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    for row_number in range(2, last_row + 1):
        worksheet.cell(
            row=row_number,
            column=1,
        ).number_format = "dd.mm.yyyy"

        for column_number in (2, 3, 4):
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).number_format = '#,##0.00'

    # Условная заливка отклонений.
    for row_number in range(2, last_row + 1):
        deviation_cell = worksheet.cell(
            row=row_number,
            column=4,
        )

        if deviation_cell.value is None:
            continue

        if deviation_cell.value > 0:
            deviation_cell.fill = PatternFill(
                fill_type="solid",
                fgColor="FCE4D6",
            )
        else:
            deviation_cell.fill = PatternFill(
                fill_type="solid",
                fgColor="E2F0D9",
            )

    # --------------------------------------------------------
    # ГРАФИК ФАКТ / ПРОГНОЗ
    # --------------------------------------------------------

    chart = LineChart()

    chart.title = sheet_title
    chart.style = 2
    chart.height = 14
    chart.width = 28

    # --------------------------------------------------------
    # ОСЬ Y — ОБЪЁМ
    # --------------------------------------------------------

    chart.y_axis.title = "Объем"
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.crosses = "autoZero"
    chart.y_axis.majorTickMark = "none"
    chart.y_axis.minorTickMark = "none"

    # Формат значений слева:
    # 120 000 вместо 120000
    chart.y_axis.numFmt = '#,##0'

    # --------------------------------------------------------
    # ОСЬ X — ДАТЫ
    # --------------------------------------------------------

    chart.x_axis.tickLblPos = "low"
    chart.x_axis.numFmt = "dd.mm.yyyy"
    chart.x_axis.majorTickMark = "none"
    chart.x_axis.minorTickMark = "none"

    # Показываем не каждую дату, чтобы подписи не накладывались.
    # Для месяца получится примерно одна подпись каждые 3 дня.
    chart.x_axis.tickLblSkip = 2
    chart.x_axis.tickMarkSkip = 2

    # Легенда снизу
    chart.legend.position = "b"

    # --------------------------------------------------------
    # ДАННЫЕ
    # --------------------------------------------------------

    chart_data = Reference(
        worksheet,
        min_col=2,
        max_col=3,
        min_row=1,
        max_row=last_row,
    )

    categories = Reference(
        worksheet,
        min_col=1,
        min_row=2,
        max_row=last_row,
    )

    chart.add_data(
        chart_data,
        titles_from_data=True,
    )

    chart.set_categories(categories)

    # --------------------------------------------------------
    # ЛИНИИ
    # --------------------------------------------------------

    if len(chart.series) >= 2:
        # Факт — синяя линия
        fact_series = chart.series[0]
        fact_series.graphicalProperties.line.solidFill = "4F81BD"
        fact_series.graphicalProperties.line.width = 32000
        fact_series.marker.symbol = "none"
        fact_series.smooth = False

        # Прогноз — красная линия
        forecast_series = chart.series[1]
        forecast_series.graphicalProperties.line.solidFill = "C0504D"
        forecast_series.graphicalProperties.line.width = 32000
        forecast_series.marker.symbol = "none"
        forecast_series.smooth = False

    # Не показываем подписи возле точек
    chart.dataLabels = None

    # Размещаем график справа от таблицы
    worksheet.add_chart(
        chart,
        "F2",
    )

@router.get("/health")
def health():
    return {
        "status": "ok",
        "module": "daily forecast",
    }


@router.post("/build/{segment_id}")
async def build_daily_forecast_endpoint(
    segment_id: str,
    session: AsyncSession = Depends(get_session),
):
    return await build_daily_forecast(
        segment_id=segment_id,
        session=session,
    )


@router.get("/current/{segment_id}")
async def get_current_daily_forecast_endpoint(
    segment_id: str,
    session: AsyncSession = Depends(get_session),
):
    run = await get_current_daily_forecast(
        segment_id=segment_id,
        session=session,
    )

    if run is None:
        return {
            "status": "not_found",
            "segment_id": segment_id,
            "message": "Current daily forecast not found",
        }

    ezmes_repo = EzmesRepository(session)

    latest_fact_updated_at = (
        await ezmes_repo.get_latest_update_time(
            segment_id=segment_id,
        )
    )

    result = serialize_run(run)
    result["latest_fact_updated_at"] = (
        latest_fact_updated_at
    )

    return result


@router.get("/history/{segment_id}")
async def get_daily_forecast_history_endpoint(
    segment_id: str,
    session: AsyncSession = Depends(get_session),
):
    runs = await get_daily_forecast_history(
        segment_id=segment_id,
        session=session,
    )

    return [
        serialize_run(run)
        for run in runs
    ]


@router.post("/auto-update/{segment_id}/{base_ymd}")
async def auto_update_daily_forecast_endpoint(
    segment_id: str,
    base_ymd: str,
    session: AsyncSession = Depends(get_session),
):
    if len(base_ymd) != 8 or not base_ymd.isdigit():
        raise HTTPException(
            status_code=400,
            detail=(
                "base_ymd должен быть в формате "
                "YYYYMMDD, например 20260714"
            ),
        )

    return await auto_update_daily_forecast(
        segment_id=segment_id,
        base_ymd=base_ymd,
        session=session,
    )


@router.post("/force-recalculate/{segment_id}/{base_ymd}")
async def force_recalculate_daily_forecast_endpoint(
    segment_id: str,
    base_ymd: str,
    session: AsyncSession = Depends(get_session),
):
    if len(base_ymd) != 8 or not base_ymd.isdigit():
        raise HTTPException(
            status_code=400,
            detail=(
                "base_ymd должен быть в формате "
                "YYYYMMDD, например 20260714"
            ),
        )

    return await force_recalculate_daily_forecast(
        segment_id=segment_id,
        base_ymd=base_ymd,
        session=session,
    )


@router.get("/export")
async def export_daily_forecasts_endpoint(
    session: AsyncSession = Depends(get_session),
):
    """
    Формирует один Excel-файл по всем Daily entities.

    В каждом листе:
    - Период;
    - Факт;
    - Прогноз;
    - Отклонения;
    - график Факт / Прогноз.
    """

    workbook = Workbook()

    # Удаляем автоматически созданный пустой лист.
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    exported_segments = 0

    for segment_id, sheet_name in DAILY_EXPORT_SEGMENTS:
        run = await get_current_daily_forecast(
            segment_id=segment_id,
            session=session,
        )

        if run is None:
            continue

        worksheet = workbook.create_sheet(
            title=sheet_name,
        )

        format_excel_sheet(
            worksheet=worksheet,
            sheet_title=sheet_name,
            values=run.values,
        )

        exported_segments += 1

    if exported_segments == 0:
        raise HTTPException(
            status_code=404,
            detail=(
                "Не найдено ни одного текущего "
                "Daily forecast для экспорта"
            ),
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = (
        "daily_forecast_"
        f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

    headers = {
        "Content-Disposition": (
            f'attachment; filename="{filename}"'
        )
    }

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        headers=headers,
    )